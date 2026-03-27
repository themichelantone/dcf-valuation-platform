import sqlite3
import pandas as pd
import numpy as np
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import yfinance as yf

def run_full_dcf(ticker_symbol):

    print(f"\n{'='*55}")
    print(f"RUNNING FULL DCF FOR: {ticker_symbol}")
    print(f"{'='*55}")

    # ── 1. Fetch & store from Yahoo Finance ───────────────────
    print("\n[1/5] Fetching data from Yahoo Finance...")
    ticker  = yf.Ticker(ticker_symbol)
    info    = ticker.info
    name    = info.get("longName", ticker_symbol)
    sector  = info.get("sector", "Unknown")
    industry= info.get("industry", "Unknown")
    price   = info.get("currentPrice", 0)

    income_stmt   = ticker.financials
    balance_sheet = ticker.balance_sheet
    cash_flow     = ticker.cashflow

    def get_row(df, keys):
        for k in keys:
            if k in df.index:
                return df.loc[k]
        return pd.Series(dtype=float)

    revenue   = get_row(income_stmt,   ["Total Revenue"])
    op_income = get_row(income_stmt,   ["Operating Income"])
    net_income= get_row(income_stmt,   ["Net Income"])
    ebit      = get_row(income_stmt,   ["EBIT","Operating Income"])
    ebitda    = get_row(income_stmt,   ["EBITDA","Normalized EBITDA"])
    cost_rev  = get_row(income_stmt,   ["Cost Of Revenue"])
    gross_prof= get_row(income_stmt,   ["Gross Profit"])
    da        = get_row(cash_flow,     ["Depreciation And Amortization",
                                        "Depreciation Amortization Depletion"])
    capex     = get_row(cash_flow,     ["Capital Expenditure"])
    fcf       = get_row(cash_flow,     ["Free Cash Flow"])
    op_cf     = get_row(cash_flow,     ["Operating Cash Flow"])
    cash      = get_row(balance_sheet, ["Cash And Cash Equivalents",
                                        "Cash Cash Equivalents And Short Term Investments"])
    debt      = get_row(balance_sheet, ["Total Debt"])
    shares    = get_row(balance_sheet, ["Share Issued","Ordinary Shares Number"])
    wc        = get_row(balance_sheet, ["Working Capital"])

    # ── 2. Store in database ──────────────────────────────────
    print("[2/5] Storing in database...")
    conn   = sqlite3.connect("dcf_database.db")
    cursor = conn.cursor()

    cursor.execute("""
        INSERT OR REPLACE INTO companies (ticker, company_name, sector, industry)
        VALUES (?, ?, ?, ?)
    """, (ticker_symbol, name, sector, industry))

    years = revenue.index if len(revenue) > 0 else []
    for year in years:
        def safe(s):
            try:
                v = s[year]
                return float(v) if pd.notna(v) else None
            except: return None
        capex_val = safe(capex)
        if capex_val: capex_val = abs(capex_val)
        cursor.execute("""
            INSERT OR REPLACE INTO financials (
                ticker, fiscal_year, total_revenue, cost_of_revenue,
                gross_profit, operating_income, net_income, ebit, ebitda,
                depreciation_amortization, capex, free_cash_flow,
                cash_and_short_term_investments, total_debt, shares_issued,
                working_capital, operating_cash_flow
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (ticker_symbol, str(year)[:10], safe(revenue), safe(cost_rev),
              safe(gross_prof), safe(op_income), safe(net_income), safe(ebit),
              safe(ebitda), safe(da), capex_val, safe(fcf), safe(cash),
              safe(debt), safe(shares), safe(wc), safe(op_cf)))

    conn.commit()

    # ── 3. Calculate DCF assumptions from history ─────────────
    print("[3/5] Calculating DCF...")
    valid = revenue.dropna()
    if len(valid) < 2:
        print("❌ Not enough data.")
        conn.close()
        return

    base_revenue    = float(valid.iloc[0])
    base_ebit_margin= float(op_income.dropna().iloc[0]) / base_revenue
    base_da_pct     = abs(float(da.dropna().iloc[0]))    / base_revenue
    base_capex_pct  = abs(float(capex.dropna().iloc[0])) / base_revenue
    base_cash       = abs(float(cash.dropna().iloc[0]))
    base_debt       = abs(float(debt.dropna().iloc[0]))
    base_shares     = abs(float(shares.dropna().iloc[0]))

    # Normalize margin (cap at 40% for non-semis)
    ebit_margin     = min(base_ebit_margin, 0.40)
    wacc            = 0.10
    terminal_growth = 0.03
    tax_rate        = 0.21
    nwc_pct         = 0.03
    growth_rates    = [0.10, 0.08, 0.07, 0.06, 0.05]

    # Forecast
    revenues, ufcf_list, pv_ufcf = [], [], []
    rev = base_revenue
    prev_rev = base_revenue
    for i, g in enumerate(growth_rates):
        rev      = rev * (1 + g)
        nopat    = rev * ebit_margin * (1 - tax_rate)
        d        = rev * base_da_pct
        cx       = rev * base_capex_pct
        dnwc     = (rev - prev_rev) * nwc_pct
        uf       = nopat + d - cx - dnwc
        pv       = uf / ((1 + wacc) ** (i + 1))
        revenues.append(rev)
        ufcf_list.append(uf)
        pv_ufcf.append(pv)
        prev_rev = rev

    tv    = (ufcf_list[-1] * (1 + terminal_growth)) / (wacc - terminal_growth)
    pv_tv = tv / ((1 + wacc) ** 5)
    ev    = sum(pv_ufcf) + pv_tv
    eq_val= ev + base_cash - base_debt
    price_per_share = eq_val / base_shares

    # Save valuation
    cursor.execute("""
        INSERT INTO valuations (
            ticker, valuation_date, wacc, terminal_growth,
            ebit_margin, enterprise_value, equity_value, implied_share_price
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (ticker_symbol, str(date.today()), wacc, terminal_growth,
          ebit_margin, ev, eq_val, price_per_share))
    conn.commit()
    conn.close()

    # ── 4. Sensitivity table ──────────────────────────────────
    def run_sens(w, tg):
        r, pr = base_revenue, base_revenue
        ufs = []
        for g in growth_rates:
            r   = r * (1 + g)
            uf  = r*ebit_margin*(1-tax_rate) + r*base_da_pct - r*base_capex_pct - (r-pr)*nwc_pct
            ufs.append(uf)
            pr  = r
        pv_cf = sum([ufs[i]/((1+w)**(i+1)) for i in range(5)])
        ptv   = (ufs[-1]*(1+tg)/(w-tg))/((1+w)**5)
        return round((pv_cf + ptv + base_cash - base_debt) / base_shares, 2)

    wacc_range = [0.08, 0.09, 0.10, 0.11, 0.12]
    tg_range   = [0.02, 0.025, 0.03, 0.035, 0.04]

    # ── 5. Export to Excel ────────────────────────────────────
    print("[4/5] Exporting to Excel...")
    wb  = Workbook()
    GREEN    = PatternFill("solid", fgColor="1F4E79")
    DARKGRAY = PatternFill("solid", fgColor="2E2E2E")
    LBLUE    = PatternFill("solid", fgColor="D6E4F0")
    GOLD     = PatternFill("solid", fgColor="F4B942")
    WHITE    = PatternFill("solid", fgColor="FFFFFF")

    def bc(ws, r, c, v, fill=None, font=None, fmt=None, align="center"):
        cell = ws.cell(row=r, column=c, value=v)
        if fill: cell.fill = fill
        if font: cell.font = font
        cell.alignment = Alignment(horizontal=align, vertical="center")
        s = Side(style="thin", color="CCCCCC")
        cell.border = Border(left=s, right=s, top=s, bottom=s)
        if fmt: cell.number_format = fmt
        return cell

    # Tab 1 — Summary
    ws1 = wb.active
    ws1.title = "Valuation Summary"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 22

    bc(ws1,1,1, f"{name} — DCF VALUATION SUMMARY",
       fill=GREEN, font=Font(bold=True,color="FFFFFF",size=13))
    ws1.merge_cells("A1:B1")
    bc(ws1,2,1, f"Valuation Date: {date.today()} | Market Price: ${price:,.2f}",
       fill=DARKGRAY, font=Font(color="FFFFFF",size=10))
    ws1.merge_cells("A2:B2")

    bc(ws1,4,1,"Item",  fill=DARKGRAY,font=Font(bold=True,color="FFFFFF"))
    bc(ws1,4,2,"Value", fill=DARKGRAY,font=Font(bold=True,color="FFFFFF"))

    rows = [
        ("PV of Cash Flows",         sum(pv_ufcf), '#,##0.00'),
        ("PV of Terminal Value",      pv_tv,        '#,##0.00'),
        ("Enterprise Value",          ev,            '#,##0.00'),
        ("+ Cash",                    base_cash,     '#,##0.00'),
        ("- Debt",                    base_debt,     '#,##0.00'),
        ("Equity Value",              eq_val,        '#,##0.00'),
        ("Shares Outstanding",        base_shares,   '#,##0'),
        ("Implied Share Price ($)",   price_per_share,'$#,##0.00'),
        ("Market Price ($)",          price,         '$#,##0.00'),
        ("Upside / Downside",         (price_per_share-price)/price, '0.0%'),
    ]
    for i, (label, val, fmt) in enumerate(rows):
        f = GOLD if "Implied" in label else (LBLUE if i%2==0 else WHITE)
        fn= Font(bold=True,size=11) if "Implied" in label else Font(size=10)
        bc(ws1,i+5,1,label,fill=f,font=fn,align="left")
        bc(ws1,i+5,2,val,  fill=f,font=fn,fmt=fmt)

    # Tab 2 — Forecast
    ws2 = wb.create_sheet("DCF Forecast")
    ws2.sheet_view.showGridLines = False
    for col,w in enumerate([28,14,14,14,14,14],1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    bc(ws2,1,1,f"{name} — 5 YEAR DCF FORECAST",
       fill=GREEN,font=Font(bold=True,color="FFFFFF",size=13))
    ws2.merge_cells("A1:F1")

    for i,h in enumerate(["Metric","Yr1","Yr2","Yr3","Yr4","Yr5"]):
        bc(ws2,3,i+1,h,fill=DARKGRAY,font=Font(bold=True,color="FFFFFF"))

    forecast_rows = [
        ("Revenue Growth", [f"{g*100:.0f}%" for g in growth_rates]),
        ("Revenue",        revenues),
        ("EBIT",           [r*ebit_margin for r in revenues]),
        ("NOPAT",          [r*ebit_margin*(1-tax_rate) for r in revenues]),
        ("D&A",            [r*base_da_pct for r in revenues]),
        ("Capex",          [r*base_capex_pct for r in revenues]),
        ("UFCF",           ufcf_list),
        ("PV of UFCF",     pv_ufcf),
    ]
    for ri,(label,vals) in enumerate(forecast_rows):
        f = LBLUE if ri%2==0 else WHITE
        bc(ws2,ri+4,1,label,fill=f,font=Font(size=10),align="left")
        for ci,v in enumerate(vals):
            bc(ws2,ri+4,ci+2,v,fill=f,font=Font(size=10),
               fmt='#,##0' if isinstance(v,float) else None)

    # Tab 3 — Sensitivity
    ws3 = wb.create_sheet("Sensitivity Analysis")
    ws3.sheet_view.showGridLines = False
    for col in range(1,8):
        ws3.column_dimensions[get_column_letter(col)].width = 15

    bc(ws3,1,1,f"{name} — SENSITIVITY (Implied Share Price $)",
       fill=GREEN,font=Font(bold=True,color="FFFFFF",size=13))
    ws3.merge_cells("A1:G1")
    bc(ws3,3,1,"TG \\ WACC",fill=DARKGRAY,font=Font(bold=True,color="FFFFFF"))
    for i,w in enumerate(wacc_range):
        bc(ws3,3,i+2,f"{w*100:.1f}%",fill=DARKGRAY,font=Font(bold=True,color="FFFFFF"))

    for ri,tg in enumerate(tg_range):
        bc(ws3,ri+4,1,f"{tg*100:.1f}%",fill=DARKGRAY,font=Font(bold=True,color="FFFFFF"))
        for ci,w in enumerate(wacc_range):
            p   = run_sens(w,tg)
            isb = (w==0.10 and tg==0.03)
            bc(ws3,ri+4,ci+2,p,
               fill=GOLD if isb else (LBLUE if ri%2==0 else WHITE),
               font=Font(bold=isb,size=10),fmt='$#,##0.00')

    # Save
    filename = f"{ticker_symbol}_DCF_Model.xlsx"
    wb.save(filename)

    print(f"[5/5] Done!")
    print(f"\n{'='*55}")
    print(f"✅ {name} DCF Complete")
    print(f"   Implied Share Price: ${price_per_share:,.2f}")
    print(f"   Market Price:        ${price:,.2f}")
    upside = (price_per_share - price) / price * 100
    print(f"   Upside/Downside:     {upside:+.1f}%")
    print(f"   Excel saved:         {filename}")
    print(f"{'='*55}\n")

# ── Run ───────────────────────────────────────────────────────
run_full_dcf("MSFT")  # Microsoft
run_full_dcf("AMZN")  # Amazon
run_full_dcf("GOOGL") # Google
