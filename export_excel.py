import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

# ── 1. Load Data ──────────────────────────────────────────────
df = pd.read_csv("nvda_financials.csv", comment="#", index_col=0)
df = df.T
df = df.apply(pd.to_numeric, errors="coerce")

base_revenue   = df["total_revenue"].iloc[0]
base_cash      = df["cash_and_short_term_investments"].iloc[0]
base_debt      = df["total_debt"].iloc[0]
base_shares    = df["shares_issued"].iloc[0]
base_da_pct    = df["depreciation_amortization"].iloc[0] / base_revenue
base_capex_pct = df["capex"].iloc[0] / base_revenue

# ── 2. Run DCF ────────────────────────────────────────────────
revenue_growth = [0.45, 0.30, 0.20, 0.15, 0.10]
ebit_margin    = 0.55
tax_rate       = 0.15
nwc_pct_change = 0.03
wacc           = 0.10
terminal_growth= 0.03

revenues, ufcf_list, pv_ufcf = [], [], []
rev = base_revenue
prev_rev = base_revenue

for i, g in enumerate(revenue_growth):
    rev       = rev * (1 + g)
    nopat     = rev * ebit_margin * (1 - tax_rate)
    da        = rev * base_da_pct
    capex     = rev * base_capex_pct
    delta_nwc = (rev - prev_rev) * nwc_pct_change
    ufcf      = nopat + da - capex - delta_nwc
    pv        = ufcf / ((1 + wacc) ** (i + 1))
    revenues.append(rev)
    ufcf_list.append(ufcf)
    pv_ufcf.append(pv)
    prev_rev  = rev

tv            = (ufcf_list[-1] * (1 + terminal_growth)) / (wacc - terminal_growth)
pv_tv         = tv / ((1 + wacc) ** 5)
ev            = sum(pv_ufcf) + pv_tv
equity_value  = ev + base_cash - base_debt
price_per_share = equity_value / base_shares

# ── 3. Sensitivity Table ──────────────────────────────────────
def run_dcf(w, tg):
    revs, ufs = [], []
    r = base_revenue
    pr = base_revenue
    for g in revenue_growth:
        r  = r * (1 + g)
        nopat = r * ebit_margin * (1 - tax_rate)
        uf = nopat + r*base_da_pct - r*base_capex_pct - (r-pr)*nwc_pct_change
        ufs.append(uf)
        pr = r
    pv_cf = sum([ufs[i]/((1+w)**(i+1)) for i in range(5)])
    ptv   = (ufs[-1]*(1+tg)/(w-tg)) / ((1+w)**5)
    return round((pv_cf + ptv + base_cash - base_debt) / base_shares, 2)

wacc_range = [0.08, 0.09, 0.10, 0.11, 0.12]
tg_range   = [0.02, 0.025, 0.03, 0.035, 0.04]

# ── 4. Styles ─────────────────────────────────────────────────
GREEN     = PatternFill("solid", fgColor="1F4E79")
DARKGRAY  = PatternFill("solid", fgColor="2E2E2E")
LIGHTBLUE = PatternFill("solid", fgColor="D6E4F0")
HIGHLIGHT = PatternFill("solid", fgColor="F4B942")
WHITE     = PatternFill("solid", fgColor="FFFFFF")

def hdr(bold=True, white=True, size=11):
    return Font(bold=bold, color="FFFFFF" if white else "000000", size=size)

def border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def write_cell(ws, row, col, value, fill=None, font=None, align="center", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:   c.fill = fill
    if font:   c.font = font
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = border()
    if num_fmt: c.number_format = num_fmt
    return c

# ── 5. Build Workbook ─────────────────────────────────────────
wb = Workbook()

# ── TAB 1: ASSUMPTIONS ────────────────────────────────────────
ws1 = wb.active
ws1.title = "Assumptions"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 20

write_cell(ws1, 1, 1, "NVIDIA DCF MODEL — ASSUMPTIONS",
           fill=GREEN, font=Font(bold=True, color="FFFFFF", size=13))
ws1.merge_cells("A1:B1")

headers = ["Parameter", "Value"]
for i, h in enumerate(headers):
    write_cell(ws1, 3, i+1, h, fill=DARKGRAY, font=hdr())

assumptions = [
    ("WACC",               f"{wacc*100:.1f}%"),
    ("Terminal Growth Rate",f"{terminal_growth*100:.1f}%"),
    ("EBIT Margin",        f"{ebit_margin*100:.1f}%"),
    ("Tax Rate",           f"{tax_rate*100:.1f}%"),
    ("NWC % Change",       f"{nwc_pct_change*100:.1f}%"),
    ("Forecast Years",     "5"),
    ("Base Year",          "1/31/2026"),
    ("Revenue Yr1 Growth", "45%"),
    ("Revenue Yr2 Growth", "30%"),
    ("Revenue Yr3 Growth", "20%"),
    ("Revenue Yr4 Growth", "15%"),
    ("Revenue Yr5 Growth", "10%"),
]

for i, (param, val) in enumerate(assumptions):
    fill = LIGHTBLUE if i % 2 == 0 else WHITE
    write_cell(ws1, i+4, 1, param, fill=fill,
               font=Font(size=10), align="left")
    write_cell(ws1, i+4, 2, val,   fill=fill,
               font=Font(bold=True, size=10))

# ── TAB 2: DCF FORECAST ───────────────────────────────────────
ws2 = wb.create_sheet("DCF Forecast")
ws2.sheet_view.showGridLines = False

for col, w in enumerate([28,15,15,15,15,15], 1):
    ws2.column_dimensions[get_column_letter(col)].width = w

write_cell(ws2, 1, 1, "NVIDIA DCF — 5 YEAR FORECAST (in $000s)",
           fill=GREEN, font=Font(bold=True, color="FFFFFF", size=13))
ws2.merge_cells("A1:F1")

col_headers = ["Metric", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
for i, h in enumerate(col_headers):
    write_cell(ws2, 3, i+1, h, fill=DARKGRAY, font=hdr())

years_label = ["2027E", "2028E", "2029E", "2030E", "2031E"]
write_cell(ws2, 4, 1, "Fiscal Year", fill=LIGHTBLUE, font=Font(bold=True, size=10))
for i, y in enumerate(years_label):
    write_cell(ws2, 4, i+2, y, fill=LIGHTBLUE, font=Font(bold=True, size=10))

rows = [
    ("Revenue Growth",  [f"{g*100:.0f}%" for g in revenue_growth]),
    ("Revenue ($000s)", revenues),
    ("EBIT",            [r * ebit_margin for r in revenues]),
    ("NOPAT",           [r * ebit_margin * (1-tax_rate) for r in revenues]),
    ("D&A",             [r * base_da_pct for r in revenues]),
    ("Capex",           [r * base_capex_pct for r in revenues]),
    ("UFCF",            ufcf_list),
    ("PV of UFCF",      pv_ufcf),
]

for ri, (label, vals) in enumerate(rows):
    row_num = ri + 5
    fill = LIGHTBLUE if ri % 2 == 0 else WHITE
    write_cell(ws2, row_num, 1, label, fill=fill,
               font=Font(size=10), align="left")
    for ci, v in enumerate(vals):
        fmt = '#,##0' if isinstance(v, float) and v > 1 else None
        write_cell(ws2, row_num, ci+2, v, fill=fill,
                   font=Font(size=10), num_fmt=fmt)

# ── TAB 3: VALUATION SUMMARY ──────────────────────────────────
ws3 = wb.create_sheet("Valuation Summary")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 20

write_cell(ws3, 1, 1, "NVIDIA DCF — VALUATION SUMMARY",
           fill=GREEN, font=Font(bold=True, color="FFFFFF", size=13))
ws3.merge_cells("A1:B1")

write_cell(ws3, 3, 1, "Item",  fill=DARKGRAY, font=hdr())
write_cell(ws3, 3, 2, "Value", fill=DARKGRAY, font=hdr())

summary = [
    ("PV of Cash Flows ($000s)",  sum(pv_ufcf)),
    ("PV of Terminal Value ($000s)", pv_tv),
    ("Enterprise Value ($000s)",  ev),
    ("+ Cash ($000s)",            base_cash),
    ("- Debt ($000s)",            base_debt),
    ("Equity Value ($000s)",      equity_value),
    ("Shares Outstanding (000s)", base_shares),
    ("Implied Share Price ($)",   price_per_share),
]

for i, (label, val) in enumerate(summary):
    fill = HIGHLIGHT if label.startswith("Implied") else (LIGHTBLUE if i%2==0 else WHITE)
    font = Font(bold=True, size=11) if label.startswith("Implied") else Font(size=10)
    write_cell(ws3, i+4, 1, label, fill=fill, font=font, align="left")
    fmt = '$#,##0.00' if label.startswith("Implied") else '#,##0'
    write_cell(ws3, i+4, 2, val,   fill=fill, font=font, num_fmt=fmt)

# ── TAB 4: SENSITIVITY ────────────────────────────────────────
ws4 = wb.create_sheet("Sensitivity Analysis")
ws4.sheet_view.showGridLines = False

for col in range(1, 8):
    ws4.column_dimensions[get_column_letter(col)].width = 16

write_cell(ws4, 1, 1, "NVIDIA DCF — SENSITIVITY ANALYSIS (Implied Share Price $)",
           fill=GREEN, font=Font(bold=True, color="FFFFFF", size=13))
ws4.merge_cells("A1:G1")

write_cell(ws4, 3, 1, "TG \\ WACC", fill=DARKGRAY, font=hdr())
for i, w in enumerate(wacc_range):
    write_cell(ws4, 3, i+2, f"{w*100:.1f}%", fill=DARKGRAY, font=hdr())

for ri, tg in enumerate(tg_range):
    write_cell(ws4, ri+4, 1, f"{tg*100:.1f}%",
               fill=DARKGRAY, font=hdr())
    for ci, w in enumerate(wacc_range):
        price = run_dcf(w, tg)
        is_base = (w == 0.10 and tg == 0.03)
        fill = HIGHLIGHT if is_base else (LIGHTBLUE if ri%2==0 else WHITE)
        font = Font(bold=True, size=10) if is_base else Font(size=10)
        write_cell(ws4, ri+4, ci+2, price, fill=fill,
                   font=font, num_fmt='$#,##0.00')

write_cell(ws4, 10, 1, "★ = Base Case (WACC 10%, TG 3%)",
           fill=WHITE, font=Font(bold=True, color="F4B942", size=10))
ws4.merge_cells("A10:G10")

# ── 6. Save ───────────────────────────────────────────────────
output_path = "NVDA_DCF_Model.xlsx"
wb.save(output_path)
print("=" * 60)
print("✅ Excel model saved successfully!")
print(f"   File: {output_path}")
print("   Tabs: Assumptions | DCF Forecast | Valuation Summary | Sensitivity Analysis")
print("=" * 60)